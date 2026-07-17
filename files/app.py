import os
import uuid
import qrcode
import socket
import openpyxl
from io import BytesIO
from PIL import Image, ImageDraw, ImageFilter
from functools import wraps
from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, send_file, abort)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY']                     = os.environ.get('SECRET_KEY', 'dev-secret-change-in-prod')
app.config['SQLALCHEMY_DATABASE_URI']        = 'sqlite:///saas.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER']                  = 'static/uploads'
app.config['MAX_CONTENT_LENGTH']             = 16 * 1024 * 1024

ALLOWED_EXT = {'png', 'jpg', 'jpeg', 'webp', 'gif'}
db = SQLAlchemy(app)


# ══════════════════════════════════════════════
#  MODELLER
# ══════════════════════════════════════════════

class Tenant(db.Model):
    id              = db.Column(db.Integer, primary_key=True)
    slug            = db.Column(db.String(60), unique=True, nullable=False)
    restoran_adi    = db.Column(db.String(120), default='Yeni Restoran')
    restoran_adi_en = db.Column(db.String(120), default='')
    logo            = db.Column(db.String(255), default='')
    banner          = db.Column(db.String(255), default='')
    whatsapp        = db.Column(db.String(60),  default='')
    instagram       = db.Column(db.String(160), default='')
    konum_url       = db.Column(db.String(255), default='')
    splash_text     = db.Column(db.String(200), default='Lezzet menümüze hoş geldiniz')
    aktif           = db.Column(db.Boolean, default=True)
    paket           = db.Column(db.String(20), default='temel')
    created_at      = db.Column(db.DateTime, default=db.func.now())
    view_count      = db.Column(db.Integer, default=0)
    lisans_bitis    = db.Column(db.DateTime, nullable=True)
    restoran_kodu   = db.Column(db.String(20), unique=True, nullable=True)  # QRM-0001
    musteri_id      = db.Column(db.Integer, db.ForeignKey('musteri.id'), nullable=True)
    tema            = db.Column(db.String(20), default='amber')  # amber / zeytin / gece
    # Lisans (tenant seviyesinde)
    lisans_tipi     = db.Column(db.String(20), default='yillik')
    odeme_tipi      = db.Column(db.String(20), default='nakit')
    ucret           = db.Column(db.Float, default=0)
    odendi_mi       = db.Column(db.Boolean, default=False)
    sozlesme_tarihi = db.Column(db.DateTime, nullable=True)
    son_iletisim    = db.Column(db.DateTime, nullable=True)
    iletisim_notu   = db.Column(db.Text, default='')
    kategoriler     = db.relationship('Kategori',  backref='tenant', lazy=True, cascade='all, delete-orphan')
    urunler         = db.relationship('Urun',       backref='tenant', lazy=True, cascade='all, delete-orphan')
    qrcodes         = db.relationship('QRCodeItem', backref='tenant', lazy=True, cascade='all, delete-orphan')
    kullanicilar    = db.relationship('Kullanici',  backref='tenant', lazy=True, cascade='all, delete-orphan')


class Kullanici(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    tenant_id     = db.Column(db.Integer, db.ForeignKey('tenant.id'), nullable=False)
    username      = db.Column(db.String(60), nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_superuser  = db.Column(db.Boolean, default=False)
    __table_args__ = (db.UniqueConstraint('tenant_id', 'username'),)

    def set_password(self, pw):  self.password_hash = generate_password_hash(pw)
    def check_password(self, pw): return check_password_hash(self.password_hash, pw)


class SuperAdmin(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(60), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)

    def set_password(self, pw):  self.password_hash = generate_password_hash(pw)
    def check_password(self, pw): return check_password_hash(self.password_hash, pw)


class Kategori(db.Model):
    id        = db.Column(db.Integer, primary_key=True)
    tenant_id = db.Column(db.Integer, db.ForeignKey('tenant.id'), nullable=False)
    isim      = db.Column(db.String(80), nullable=False)
    isim_en   = db.Column(db.String(80), default='')
    resim     = db.Column(db.String(255), default='')
    banner    = db.Column(db.String(255), default='')
    durum     = db.Column(db.Boolean, default=True)
    sira      = db.Column(db.Integer, default=0)
    urunler   = db.relationship('Urun', backref='kategori', lazy=True, cascade='all, delete-orphan')


class Urun(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    tenant_id     = db.Column(db.Integer, db.ForeignKey('tenant.id'),   nullable=False)
    kategori_id   = db.Column(db.Integer, db.ForeignKey('kategori.id'), nullable=False)
    isim          = db.Column(db.String(120), nullable=False)
    isim_en       = db.Column(db.String(120), default='')
    fiyat         = db.Column(db.Float, default=0)
    aciklama      = db.Column(db.Text, default='')
    aciklama_en   = db.Column(db.Text, default='')
    resim         = db.Column(db.String(255), default='')
    durum         = db.Column(db.Boolean, default=True)
    one_cikan     = db.Column(db.Boolean, default=False)
    badge_yeni    = db.Column(db.Boolean, default=False)
    badge_populer = db.Column(db.Boolean, default=False)
    badge_acili   = db.Column(db.Boolean, default=False)
    sira          = db.Column(db.Integer, default=0)


class QRCodeItem(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    tenant_id   = db.Column(db.Integer, db.ForeignKey('tenant.id'), nullable=False)
    isim        = db.Column(db.String(100), nullable=False)
    hedef_url   = db.Column(db.String(255), nullable=False)
    dosya       = db.Column(db.String(255), nullable=False)
    kilitli     = db.Column(db.Boolean, default=True)
    silme_hazir = db.Column(db.Boolean, default=False)
    renk_on     = db.Column(db.String(20), default='#000000')
    renk_arka   = db.Column(db.String(20), default='#ffffff')
    logo_var    = db.Column(db.Boolean, default=False)


class Musteri(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    # Kimlik
    musteri_kodu  = db.Column(db.String(20), unique=True, nullable=False)  # MUS-0001
    ad_soyad      = db.Column(db.String(120), nullable=False)
    telefon       = db.Column(db.String(30),  default='')
    email         = db.Column(db.String(120), default='')
    tc_vkn        = db.Column(db.String(20),  default='')
    sehir         = db.Column(db.String(60),  default='')
    ilce          = db.Column(db.String(60),  default='')
    notlar        = db.Column(db.Text,        default='')
    created_at    = db.Column(db.DateTime,    default=db.func.now())
    # Restoranlar bu müşteriye backref ile bağlı
    restoranlar   = db.relationship('Tenant', backref='musteri', lazy=True)


# ══════════════════════════════════════════════
#  VERİTABANI BAŞLATMA
# ══════════════════════════════════════════════
with app.app_context():
    db.create_all()
    if not SuperAdmin.query.first():
        sa = SuperAdmin(username='superadmin')
        sa.set_password('superadmin123')
        db.session.add(sa)
        db.session.commit()


# ══════════════════════════════════════════════
#  YARDIMCILAR
# ══════════════════════════════════════════════
def upload_dir(slug, sub=''):
    path = os.path.join(app.config['UPLOAD_FOLDER'], slug, sub) if sub \
           else os.path.join(app.config['UPLOAD_FOLDER'], slug)
    os.makedirs(path, exist_ok=True)
    return path


def allowed(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def save_img(file, slug, sub):
    if not file or not file.filename:
        return None
    if not allowed(file.filename):
        return None
    ext  = secure_filename(file.filename).rsplit('.', 1)[1].lower()
    name = f'{uuid.uuid4().hex}.{ext}'
    file.save(os.path.join(upload_dir(slug, sub), name))
    return name


def musteri_kodu_uret():
    son = Musteri.query.order_by(Musteri.id.desc()).first()
    num = (son.id + 1) if son else 1
    return f'MUS-{num:04d}'

def restoran_kodu_uret():
    son = Tenant.query.order_by(Tenant.id.desc()).first()
    num = (son.id + 1) if son else 1
    return f'QRM-{num:04d}'

def clean(v, d=''):  return (v or d).strip()
def parse_price(v):
    try:    return float(str(v).replace(',', '.').strip())
    except: return None

def get_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80)); ip = s.getsockname()[0]; s.close(); return ip
    except: return '127.0.0.1'

def ok(m):  flash(m, 'success')
def err(m): flash(m, 'error')


# ── Dekoratörler ─────────────────────────────
def sa_required(f):
    @wraps(f)
    def w(*a, **kw):
        if not session.get('sa_id'):
            return redirect(url_for('sa_login'))
        return f(*a, **kw)
    return w


def login_required(f):
    @wraps(f)
    def w(*a, **kw):
        slug   = kw.get('slug')
        tenant = Tenant.query.filter_by(slug=slug, aktif=True).first_or_404()
        uid    = session.get(f't_{slug}')
        if not uid:
            return redirect(url_for('t_login', slug=slug))
        user = Kullanici.query.filter_by(id=uid, tenant_id=tenant.id).first()
        if not user:
            session.pop(f't_{slug}', None)
            return redirect(url_for('t_login', slug=slug))
        return f(*a, **kw, tenant=tenant, me=user)
    return w


@app.context_processor
def ctx():
    sid = session.get('sa_id')
    return {'sa': SuperAdmin.query.get(sid) if sid else None}


# ══════════════════════════════════════════════
#  QR ÜRETİCİ
# ══════════════════════════════════════════════
def make_qr(url, fg='#000000', bg='#ffffff', logo=None):
    def h(c):
        c = c.lstrip('#')
        return tuple(int(c[i:i+2], 16) for i in (0, 2, 4))

    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=14, border=3)
    qr.add_data(url); qr.make(fit=True)
    img = qr.make_image(fill_color=h(fg), back_color=h(bg)).convert('RGBA')

    if logo and os.path.exists(logo):
        qw, qh = img.size
        area   = int(min(qw, qh) * 0.22)
        pad    = int(area * 0.18)
        csz    = area + pad * 2

        sh = Image.new('RGBA', (csz+8, csz+8), (0,0,0,0))
        ImageDraw.Draw(sh).ellipse((4,4,csz+4,csz+4), fill=(0,0,0,60))
        sh = sh.filter(ImageFilter.GaussianBlur(4))

        ci = Image.new('RGBA', (csz, csz), (0,0,0,0))
        ImageDraw.Draw(ci).ellipse((0,0,csz-1,csz-1), fill=(255,255,255,255))

        lo = Image.open(logo).convert('RGBA').resize((area, area), Image.LANCZOS)
        mk = Image.new('L', (area, area), 0)
        ImageDraw.Draw(mk).ellipse((0,0,area,area), fill=255)
        lr = Image.new('RGBA', (area, area), (0,0,0,0))
        lr.paste(lo, (0,0), mk); ci.paste(lr, (pad,pad), lr)

        px, py = (qw-csz)//2, (qh-csz)//2
        img.paste(sh, (px-4, py-4), sh)
        img.paste(ci, (px, py), ci)
    return img


# ══════════════════════════════════════════════
#  SÜPER ADMİN
# ══════════════════════════════════════════════
@app.route('/')
def root(): return redirect(url_for('sa_login'))


@app.route('/superadmin/login', methods=['GET','POST'])
def sa_login():
    if request.method == 'POST':
        sa = SuperAdmin.query.filter_by(username=clean(request.form.get('username'))).first()
        if sa and sa.check_password(request.form.get('password','')):
            session['sa_id'] = sa.id
            return redirect(url_for('sa_panel'))
        err('Hatalı giriş.')
    return render_template('superadmin/login.html')


@app.route('/superadmin/logout')
def sa_logout():
    session.pop('sa_id', None)
    return redirect(url_for('sa_login'))


@app.route('/superadmin')
@sa_required
def sa_panel():
    from datetime import datetime, timedelta
    now = datetime.now()
    tenants  = Tenant.query.order_by(Tenant.id.desc()).all()
    musteriler = Musteri.query.order_by(Musteri.id.desc()).all()

    for t in tenants:
        if t.lisans_bitis and t.lisans_bitis < now and t.aktif:
            t.aktif = False
    db.session.commit()

    yaklasan      = [t for t in tenants if t.lisans_bitis and now < t.lisans_bitis < now + timedelta(days=30)]
    suresi_dolmus = [t for t in tenants if t.lisans_bitis and t.lisans_bitis < now]

    stats = dict(
        toplam_musteri=len(musteriler),
        toplam=len(tenants),
        aktif=sum(1 for t in tenants if t.aktif),
        pro=sum(1 for t in tenants if t.paket=='pro'),
        kurumsal=sum(1 for t in tenants if t.paket=='kurumsal'),
        toplam_gorunum=sum(t.view_count or 0 for t in tenants),
        toplam_gelir=sum(t.ucret or 0 for t in tenants),
        yaklasan_lisans=len(yaklasan),
        suresi_dolmus=len(suresi_dolmus),
    )
    return render_template('superadmin/dashboard.html',
        tenants=tenants, musteriler=musteriler, stats=stats,
        yaklasan=yaklasan, suresi_dolmus=suresi_dolmus, now=now)


@app.route('/superadmin/ekle', methods=['POST'])
@sa_required
def sa_ekle():
    slug  = clean(request.form.get('slug','')).lower().replace(' ','-')
    isim  = clean(request.form.get('restoran_adi','Yeni Restoran'))
    pw    = request.form.get('password','admin123')
    paket = request.form.get('paket','temel')

    if not slug:
        err('Slug zorunludur.'); return redirect(url_for('sa_panel'))
    if Tenant.query.filter_by(slug=slug).first():
        err(f'"{slug}" zaten kullanımda.'); return redirect(url_for('sa_panel'))

    t = Tenant(slug=slug, restoran_adi=isim, paket=paket)
    db.session.add(t); db.session.flush()

    u = Kullanici(tenant_id=t.id, username='admin', is_superuser=True)
    u.set_password(pw); db.session.add(u)
    db.session.commit()

    for sub in ['kategoriler','urunler','qrcodes','ayarlar']:
        upload_dir(slug, sub)

    ok(f'"{isim}" oluşturuldu — URL: /r/{slug}/  |  Şifre: {pw}')
    return redirect(url_for('sa_panel'))


@app.route('/superadmin/durum/<int:tid>')
@sa_required
def sa_durum(tid):
    t = Tenant.query.get_or_404(tid)
    t.aktif = not t.aktif; db.session.commit()
    ok('Durum güncellendi.'); return redirect(url_for('sa_panel'))


@app.route('/superadmin/sil/<int:tid>', methods=['POST'])
@sa_required
def sa_sil(tid):
    t = Tenant.query.get_or_404(tid)
    db.session.delete(t); db.session.commit()
    ok(f'"{t.restoran_adi}" silindi.'); return redirect(url_for('sa_panel'))


@app.route('/superadmin/paket/<int:tid>', methods=['POST'])
@sa_required
def sa_paket(tid):
    t = Tenant.query.get_or_404(tid)
    t.paket = request.form.get('paket','temel'); db.session.commit()
    ok('Paket güncellendi.'); return redirect(url_for('sa_panel'))


# ══════════════════════════════════════════════
#  MÜŞTERI MENÜSÜ (herkese açık)
# ══════════════════════════════════════════════
@app.route('/r/<slug>/')
def menu(slug):
    tenant = Tenant.query.filter_by(slug=slug, aktif=True).first_or_404()
    tenant.view_count = (tenant.view_count or 0) + 1
    db.session.commit()
    kats   = Kategori.query.filter_by(tenant_id=tenant.id, durum=True)\
                           .order_by(Kategori.sira, Kategori.id).all()
    result = []
    for k in kats:
        aktif = [u for u in k.urunler if u.durum]
        if aktif:
            k.aktif_urunler = aktif
            result.append(k)
    return render_template('menu/index.html', tenant=tenant, kategoriler=result)


# ══════════════════════════════════════════════
#  TENANT LOGIN/LOGOUT
# ══════════════════════════════════════════════
@app.route('/r/<slug>/admin/login', methods=['GET','POST'])
def t_login(slug):
    tenant = Tenant.query.filter_by(slug=slug, aktif=True).first_or_404()
    if request.method == 'POST':
        u = Kullanici.query.filter_by(
            tenant_id=tenant.id, username=clean(request.form.get('username'))).first()
        if u and u.check_password(request.form.get('password','')):
            session[f't_{slug}'] = u.id
            ok('Giriş başarılı.')
            return redirect(url_for('t_admin', slug=slug))
        err('Hatalı giriş.')
    return render_template('tenant/login.html', tenant=tenant)


@app.route('/r/<slug>/admin/logout')
def t_logout(slug):
    session.pop(f't_{slug}', None)
    return redirect(url_for('t_login', slug=slug))


# ══════════════════════════════════════════════
#  TENANT ADMİN PANELİ
# ══════════════════════════════════════════════
@app.route('/r/<slug>/admin')
@login_required
def t_admin(slug, tenant, me):
    kats  = Kategori.query.filter_by(tenant_id=tenant.id)\
                          .order_by(Kategori.sira, Kategori.id.desc()).all()
    uruns = Urun.query.filter_by(tenant_id=tenant.id).order_by(Urun.sira, Urun.id.desc()).all()
    users = Kullanici.query.filter_by(tenant_id=tenant.id).all()
    qrs   = QRCodeItem.query.filter_by(tenant_id=tenant.id).order_by(QRCodeItem.id.desc()).all()
    port  = request.host.split(':')[1] if ':' in request.host else '5000'
    local = f'http://{get_ip()}:{port}/r/{slug}/'
    return render_template('tenant/admin.html',
        tenant=tenant, me=me,
        kategoriler=kats, urunler=uruns,
        kullanicilar=users, qrcodes=qrs, local_url=local)


# ── Ayarlar ──
@app.route('/r/<slug>/admin/ayarlar', methods=['POST'])
@login_required
def t_ayarlar(slug, tenant, me):
    tenant.restoran_adi    = clean(request.form.get('restoran_adi'), tenant.restoran_adi)
    tenant.restoran_adi_en = clean(request.form.get('restoran_adi_en'))
    tenant.whatsapp        = clean(request.form.get('whatsapp'))
    tenant.instagram       = clean(request.form.get('instagram'))
    tenant.konum_url       = clean(request.form.get('konum_url'))
    tenant.splash_text     = clean(request.form.get('splash_text'), tenant.splash_text)
    tema = request.form.get('tema', '').strip()
    if tema in ('amber', 'zeytin', 'gece'):
        tenant.tema = tema
    logo   = save_img(request.files.get('logo'),   slug, 'ayarlar')
    banner = save_img(request.files.get('banner'), slug, 'ayarlar')
    if logo:   tenant.logo   = logo
    if banner: tenant.banner = banner
    db.session.commit(); ok('Ayarlar kaydedildi.')
    return redirect(url_for('t_admin', slug=slug) + '#ayarlar')


# ── Kategoriler ──
@app.route('/r/<slug>/admin/kat_ekle', methods=['POST'])
@login_required
def t_kat_ekle(slug, tenant, me):
    isim = clean(request.form.get('isim'))
    if not isim:
        err('Ad zorunludur.')
    elif Kategori.query.filter_by(tenant_id=tenant.id, isim=isim).first():
        err('Bu isim zaten var.')
    else:
        resim  = save_img(request.files.get('resim'),  slug, 'kategoriler')
        banner = save_img(request.files.get('banner'), slug, 'kategoriler')
        db.session.add(Kategori(tenant_id=tenant.id, isim=isim,
            isim_en=clean(request.form.get('isim_en'), isim),
            resim=resim or '', banner=banner or '', durum=True))
        db.session.commit(); ok('Kategori eklendi.')
    return redirect(url_for('t_admin', slug=slug) + '#kategoriler')


@app.route('/r/<slug>/admin/kat_duzenle/<int:kid>', methods=['POST'])
@login_required
def t_kat_duzenle(slug, tenant, me, kid):
    k = Kategori.query.filter_by(id=kid, tenant_id=tenant.id).first_or_404()
    k.isim = clean(request.form.get('isim'), k.isim)
    k.isim_en = clean(request.form.get('isim_en'))
    yeni        = save_img(request.files.get('resim'),  slug, 'kategoriler')
    yeni_banner = save_img(request.files.get('banner'), slug, 'kategoriler')
    if yeni:        k.resim  = yeni
    if yeni_banner: k.banner = yeni_banner
    db.session.commit(); ok('Güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#kategoriler')


@app.route('/r/<slug>/admin/kat_durum/<int:kid>')
@login_required
def t_kat_durum(slug, tenant, me, kid):
    k = Kategori.query.filter_by(id=kid, tenant_id=tenant.id).first_or_404()
    k.durum = not k.durum; db.session.commit(); ok('Güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#kategoriler')


@app.route('/r/<slug>/admin/kat_sil/<int:kid>')
@login_required
def t_kat_sil(slug, tenant, me, kid):
    k = Kategori.query.filter_by(id=kid, tenant_id=tenant.id).first_or_404()
    db.session.delete(k); db.session.commit(); ok('Silindi.')
    return redirect(url_for('t_admin', slug=slug) + '#kategoriler')


# ── Ürünler ──
@app.route('/r/<slug>/admin/urun_ekle', methods=['POST'])
@login_required
def t_urun_ekle(slug, tenant, me):
    isim  = clean(request.form.get('isim'))
    fiyat = parse_price(request.form.get('fiyat'))
    kid   = request.form.get('kategori_id')
    if not isim or fiyat is None or not kid:
        err('Zorunlu alanları doldur.')
        return redirect(url_for('t_admin', slug=slug) + '#urunler')
    kat = Kategori.query.filter_by(id=int(kid), tenant_id=tenant.id).first()
    if not kat:
        err('Geçersiz kategori.')
        return redirect(url_for('t_admin', slug=slug) + '#urunler')
    resim = save_img(request.files.get('resim'), slug, 'urunler')
    db.session.add(Urun(
        tenant_id=tenant.id, kategori_id=kat.id,
        isim=isim, isim_en=clean(request.form.get('isim_en'), isim),
        fiyat=fiyat,
        aciklama=clean(request.form.get('aciklama')),
        aciklama_en=clean(request.form.get('aciklama_en')),
        resim=resim or '', durum=True,
        one_cikan=bool(request.form.get('one_cikan')),
        badge_yeni=bool(request.form.get('badge_yeni')),
        badge_populer=bool(request.form.get('badge_populer')),
        badge_acili=bool(request.form.get('badge_acili')),
    ))
    db.session.commit(); ok('Ürün eklendi.')
    return redirect(url_for('t_admin', slug=slug) + '#urunler')


@app.route('/r/<slug>/admin/urun_duzenle/<int:uid>', methods=['POST'])
@login_required
def t_urun_duzenle(slug, tenant, me, uid):
    u = Urun.query.filter_by(id=uid, tenant_id=tenant.id).first_or_404()
    fiyat = parse_price(request.form.get('fiyat'))
    kid   = request.form.get('kategori_id')
    u.isim        = clean(request.form.get('isim'), u.isim)
    u.isim_en     = clean(request.form.get('isim_en'))
    u.aciklama    = clean(request.form.get('aciklama'))
    u.aciklama_en = clean(request.form.get('aciklama_en'))
    u.one_cikan   = bool(request.form.get('one_cikan'))
    u.badge_yeni  = bool(request.form.get('badge_yeni'))
    u.badge_populer = bool(request.form.get('badge_populer'))
    u.badge_acili = bool(request.form.get('badge_acili'))
    if fiyat is not None: u.fiyat = fiyat
    if kid:
        k = Kategori.query.filter_by(id=int(kid), tenant_id=tenant.id).first()
        if k: u.kategori_id = k.id
    yeni = save_img(request.files.get('resim'), slug, 'urunler')
    if yeni: u.resim = yeni
    db.session.commit(); ok('Güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#urunler')


@app.route('/r/<slug>/admin/urun_durum/<int:uid>')
@login_required
def t_urun_durum(slug, tenant, me, uid):
    u = Urun.query.filter_by(id=uid, tenant_id=tenant.id).first_or_404()
    u.durum = not u.durum; db.session.commit(); ok('Güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#urunler')


@app.route('/r/<slug>/admin/urun_one/<int:uid>')
@login_required
def t_urun_one(slug, tenant, me, uid):
    u = Urun.query.filter_by(id=uid, tenant_id=tenant.id).first_or_404()
    u.one_cikan = not u.one_cikan; db.session.commit(); ok('Güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#urunler')


@app.route('/r/<slug>/admin/urun_sil/<int:uid>')
@login_required
def t_urun_sil(slug, tenant, me, uid):
    u = Urun.query.filter_by(id=uid, tenant_id=tenant.id).first_or_404()
    db.session.delete(u); db.session.commit(); ok('Silindi.')
    return redirect(url_for('t_admin', slug=slug) + '#urunler')


# ── Excel Aktarım ──
@app.route('/r/<slug>/admin/excel_sablon')
@login_required
def t_excel_sablon(slug, tenant, me):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Urunler'
    ws.append(['Kategori Adı','Ürün Adı (TR)','Ürün Adı (EN)',
               'Fiyat','Açıklama (TR)','Açıklama (EN)',
               'Öne Çıkan(1/0)','Yeni(1/0)','Popüler(1/0)','Acılı(1/0)'])
    ws.append(['Pizzalar','Margarita','Margherita','120','Klasik pizza','Classic pizza','0','1','0','0'])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='urun_sablon.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/r/<slug>/admin/excel_yukle', methods=['POST'])
@login_required
def t_excel_yukle(slug, tenant, me):
    from flask import jsonify
    f = request.files.get('excel_file')
    if not f or not f.filename.endswith(('.xlsx','.xls')):
        return jsonify({'ok': False, 'mesaj': 'Geçerli bir .xlsx dosyası seç.', 'hatalar': []})
    try:
        wb = openpyxl.load_workbook(f, data_only=True); ws = wb.active
    except Exception:
        return jsonify({'ok': False, 'mesaj': 'Dosya okunamadı.', 'hatalar': []})

    eklenen, hatalar = 0, []
    for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row): continue
        try:
            kat_isim = str(row[0] or '').strip()
            isim     = str(row[1] or '').strip()
            if not kat_isim or not isim:
                hatalar.append(f'Satır {rn}: Kategori/ürün adı boş.'); continue
            fiyat = parse_price(row[3])
            if fiyat is None:
                hatalar.append(f'Satır {rn}: Geçersiz fiyat.'); continue
            kat = Kategori.query.filter_by(tenant_id=tenant.id, isim=kat_isim).first()
            if not kat:
                kat = Kategori(tenant_id=tenant.id, isim=kat_isim, isim_en='', resim='', durum=True)
                db.session.add(kat); db.session.flush()
            db.session.add(Urun(
                tenant_id=tenant.id, kategori_id=kat.id,
                isim=isim, isim_en=str(row[2] or isim).strip(),
                fiyat=fiyat,
                aciklama=str(row[4] or '').strip(),
                aciklama_en=str(row[5] or '').strip(),
                resim='', durum=True,
                one_cikan=str(row[6] or '0').strip()=='1',
                badge_yeni=str(row[7] or '0').strip()=='1',
                badge_populer=str(row[8] or '0').strip()=='1',
                badge_acili=str(row[9] or '0').strip()=='1',
            ))
            eklenen += 1
        except Exception as e:
            hatalar.append(f'Satır {rn}: {e}')
    db.session.commit()

    if eklenen and not hatalar:
        return jsonify({'ok': True, 'mesaj': f'{eklenen} ürün başarıyla aktarıldı!', 'hatalar': []})
    elif eklenen and hatalar:
        return jsonify({'ok': True, 'mesaj': f'{eklenen} ürün aktarıldı, bazı satırlarda hata var.', 'hatalar': hatalar[:10]})
    else:
        return jsonify({'ok': False, 'mesaj': 'Hiç ürün aktarılamadı.', 'hatalar': hatalar[:10]})


# ── Kullanıcılar ──
@app.route('/r/<slug>/admin/kullanici_ekle', methods=['POST'])
@login_required
def t_kullanici_ekle(slug, tenant, me):
    if not me.is_superuser:
        err('Yetkisiz.'); return redirect(url_for('t_admin', slug=slug) + '#kullanicilar')
    uname = clean(request.form.get('username'))
    pw    = request.form.get('password','')
    if not uname or not pw:
        err('Ad ve şifre zorunlu.')
    elif Kullanici.query.filter_by(tenant_id=tenant.id, username=uname).first():
        err('Bu kullanıcı adı var.')
    else:
        u = Kullanici(tenant_id=tenant.id, username=uname,
                      is_superuser=bool(request.form.get('is_superuser')))
        u.set_password(pw); db.session.add(u); db.session.commit(); ok('Kullanıcı eklendi.')
    return redirect(url_for('t_admin', slug=slug) + '#kullanicilar')


@app.route('/r/<slug>/admin/kullanici_sil/<int:kid>')
@login_required
def t_kullanici_sil(slug, tenant, me, kid):
    if not me.is_superuser:
        err('Yetkisiz.'); return redirect(url_for('t_admin', slug=slug) + '#kullanicilar')
    u = Kullanici.query.filter_by(id=kid, tenant_id=tenant.id).first_or_404()
    if u.id == me.id:
        err('Kendinizi silemezsiniz.')
    else:
        db.session.delete(u); db.session.commit(); ok('Silindi.')
    return redirect(url_for('t_admin', slug=slug) + '#kullanicilar')


@app.route('/r/<slug>/admin/sifre/<int:kid>', methods=['POST'])
@login_required
def t_sifre(slug, tenant, me, kid):
    u = Kullanici.query.filter_by(id=kid, tenant_id=tenant.id).first_or_404()
    if not me.is_superuser and u.id != me.id:
        err('Yetkisiz.'); return redirect(url_for('t_admin', slug=slug) + '#kullanicilar')
    pw = request.form.get('new_password','')
    if not pw: err('Şifre boş olamaz.')
    else: u.set_password(pw); db.session.commit(); ok('Şifre güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#kullanicilar')


# ── QR Kodlar ──
@app.route('/r/<slug>/admin/qr_olustur', methods=['POST'])
@login_required
def t_qr_olustur(slug, tenant, me):
    isim  = clean(request.form.get('isim'), 'Menü QR')
    url   = clean(request.form.get('hedef_url'), f'{request.host_url}r/{slug}/')
    fg    = clean(request.form.get('renk_on'),   '#000000')
    bg    = clean(request.form.get('renk_arka'), '#ffffff')
    logo_p = None
    if request.form.get('logo_ekle') and tenant.logo:
        for sub in ('ayarlar','kategoriler'):
            c = os.path.join(app.config['UPLOAD_FOLDER'], slug, sub, tenant.logo)
            if os.path.exists(c): logo_p = c; break
    fname = f'{uuid.uuid4().hex}.png'
    path  = os.path.join(upload_dir(slug, 'qrcodes'), fname)
    make_qr(url, fg, bg, logo_p).save(path)
    db.session.add(QRCodeItem(
        tenant_id=tenant.id, isim=isim, hedef_url=url, dosya=fname,
        kilitli=True, silme_hazir=False, renk_on=fg, renk_arka=bg,
        logo_var=(logo_p is not None)))
    db.session.commit(); ok('QR oluşturuldu.')
    return redirect(url_for('t_admin', slug=slug) + '#qrcodes')


@app.route('/r/<slug>/admin/qr_kilit/<int:qid>')
@login_required
def t_qr_kilit(slug, tenant, me, qid):
    q = QRCodeItem.query.filter_by(id=qid, tenant_id=tenant.id).first_or_404()
    q.kilitli = not q.kilitli
    if q.kilitli: q.silme_hazir = False
    db.session.commit(); ok('Kilit güncellendi.')
    return redirect(url_for('t_admin', slug=slug) + '#qrcodes')


@app.route('/r/<slug>/admin/qr_hazirla/<int:qid>')
@login_required
def t_qr_hazirla(slug, tenant, me, qid):
    q = QRCodeItem.query.filter_by(id=qid, tenant_id=tenant.id).first_or_404()
    if q.kilitli: err('Önce kilidi aç.')
    else: q.silme_hazir = True; db.session.commit(); ok('Adım 1 aktif.')
    return redirect(url_for('t_admin', slug=slug) + '#qrcodes')


@app.route('/r/<slug>/admin/qr_iptal/<int:qid>')
@login_required
def t_qr_iptal(slug, tenant, me, qid):
    q = QRCodeItem.query.filter_by(id=qid, tenant_id=tenant.id).first_or_404()
    q.silme_hazir = False; db.session.commit(); ok('İptal edildi.')
    return redirect(url_for('t_admin', slug=slug) + '#qrcodes')


@app.route('/r/<slug>/admin/qr_sil/<int:qid>')
@login_required
def t_qr_sil(slug, tenant, me, qid):
    q = QRCodeItem.query.filter_by(id=qid, tenant_id=tenant.id).first_or_404()
    if not q.kilitli and q.silme_hazir:
        fp = os.path.join(upload_dir(slug, 'qrcodes'), q.dosya)
        if os.path.exists(fp): os.remove(fp)
        db.session.delete(q); db.session.commit(); ok('QR silindi.')
    else: err('Koşullar sağlanmadı.')
    return redirect(url_for('t_admin', slug=slug) + '#qrcodes')


@app.route('/superadmin/sifre/<int:tid>', methods=['POST'])
@sa_required
def sa_sifre(tid):
    t = Tenant.query.get_or_404(tid)
    pw = request.form.get('new_password', '').strip()
    if not pw:
        err('Şifre boş olamaz.')
        return redirect(url_for('sa_panel'))
    u = Kullanici.query.filter_by(tenant_id=t.id, is_superuser=True).first()
    if not u:
        err('Admin kullanıcı bulunamadı.')
        return redirect(url_for('sa_panel'))
    u.set_password(pw)
    db.session.commit()
    ok(f'"{t.restoran_adi}" admin şifresi güncellendi.')
    return redirect(url_for('sa_panel'))


# ══════════════════════════════════════════════
#  MÜŞTERİ YÖNETİMİ
# ══════════════════════════════════════════════


@app.route('/superadmin/musteri_ekle', methods=['POST'])
@sa_required
def sa_musteri_ekle():
    ad = (request.form.get('ad_soyad') or '').strip()
    if not ad:
        err('Ad Soyad zorunludur.'); return redirect(url_for('sa_panel'))
    kod = musteri_kodu_uret()
    m = Musteri(
        musteri_kodu = kod,
        ad_soyad     = ad,
        telefon      = (request.form.get('telefon') or '').strip(),
        email        = (request.form.get('email') or '').strip(),
        tc_vkn       = (request.form.get('tc_vkn') or '').strip(),
        sehir        = (request.form.get('sehir') or '').strip(),
        ilce         = (request.form.get('ilce') or '').strip(),
        notlar       = (request.form.get('notlar') or '').strip(),
    )
    db.session.add(m)
    db.session.commit()
    ok(f'Müşteri "{ad}" oluşturuldu. Kod: {kod}')
    return redirect(url_for('sa_panel'))


@app.route('/superadmin/musteri_duzenle/<int:mid>', methods=['POST'])
@sa_required
def sa_musteri_duzenle(mid):
    m = Musteri.query.get_or_404(mid)
    m.ad_soyad = (request.form.get('ad_soyad') or m.ad_soyad).strip()
    m.telefon  = (request.form.get('telefon') or '').strip()
    m.email    = (request.form.get('email') or '').strip()
    m.tc_vkn   = (request.form.get('tc_vkn') or '').strip()
    m.sehir    = (request.form.get('sehir') or '').strip()
    m.ilce     = (request.form.get('ilce') or '').strip()
    m.notlar   = (request.form.get('notlar') or '').strip()
    db.session.commit()
    ok('Müşteri güncellendi.')
    return redirect(url_for('sa_panel'))


@app.route('/superadmin/musteri_sil/<int:mid>', methods=['POST'])
@sa_required
def sa_musteri_sil(mid):
    m = Musteri.query.get_or_404(mid)
    for t in m.restoranlar:
        t.musteri_id = None
    db.session.delete(m)
    db.session.commit()
    ok('Müşteri silindi.')
    return redirect(url_for('sa_panel'))


@app.route('/superadmin/lisans_ata/<int:mid>', methods=['POST'])
@sa_required
def sa_lisans_ata(mid):
    from datetime import datetime, timedelta
    m    = Musteri.query.get_or_404(mid)
    slug = (request.form.get('slug') or '').lower().strip().replace(' ', '-')
    isim = (request.form.get('restoran_adi') or m.ad_soyad).strip()
    pw   = request.form.get('password', 'admin123')
    paket = request.form.get('paket', 'temel')
    sure  = int(request.form.get('sure', 12) or 12)
    ucret = float(request.form.get('ucret', 0) or 0)
    odeme = request.form.get('odeme_tipi', 'nakit')
    odendi = bool(request.form.get('odendi_mi'))

    if not slug:
        err('Slug zorunludur.'); return redirect(url_for('sa_panel'))
    if Tenant.query.filter_by(slug=slug).first():
        err(f'"{slug}" zaten kullanımda.'); return redirect(url_for('sa_panel'))

    bas   = datetime.now()
    bitis = bas + timedelta(days=30 * sure)
    kod   = restoran_kodu_uret()

    t = Tenant(
        slug=slug, restoran_adi=isim, paket=paket,
        aktif=True, musteri_id=m.id,
        restoran_kodu=kod,
        lisans_bitis=bitis,
        lisans_tipi=request.form.get('lisans_tipi', 'yillik'),
        ucret=ucret, odeme_tipi=odeme, odendi_mi=odendi,
    )
    db.session.add(t); db.session.flush()

    u = Kullanici(tenant_id=t.id, username='admin', is_superuser=True)
    u.set_password(pw)
    db.session.add(u)
    db.session.commit()

    for sub in ['kategoriler', 'urunler', 'qrcodes', 'ayarlar']:
        upload_dir(slug, sub)

    ok(f'Restoran "{isim}" oluşturuldu. Kod: {kod} | {sure} ay lisans | /r/{slug}/')
    return redirect(url_for('sa_panel'))


@app.route('/superadmin/lisans_yenile/<int:tid>', methods=['POST'])
@sa_required
def sa_lisans_yenile(tid):
    from datetime import datetime, timedelta
    t    = Tenant.query.get_or_404(tid)
    sure = int(request.form.get('sure', 12) or 12)
    bas  = t.lisans_bitis if t.lisans_bitis and t.lisans_bitis > datetime.now() else datetime.now()
    t.lisans_bitis = bas + timedelta(days=30 * sure)
    t.aktif = True
    db.session.commit()
    ok(f'{t.restoran_adi} lisansı {sure} ay uzatıldı.')
    return redirect(url_for('sa_panel'))


@app.route('/superadmin/iletisim/<int:tid>', methods=['POST'])
@sa_required
def sa_iletisim(tid):
    from datetime import datetime
    t = Tenant.query.get_or_404(tid)
    t.iletisim_notu = (request.form.get('iletisim_notu') or '').strip()
    t.son_iletisim  = datetime.now()
    db.session.commit()
    ok('İletişim notu güncellendi.')
    return redirect(url_for('sa_panel'))



# ── Resim Kaldır ──
@app.route('/r/<slug>/admin/kat_resim_sil/<int:kid>/<tip>')
@login_required
def t_kat_resim_sil(slug, tenant, me, kid, tip):
    from flask import jsonify
    k = Kategori.query.filter_by(id=kid, tenant_id=tenant.id).first_or_404()
    if tip == 'resim':
        if k.resim:
            try: os.remove(os.path.join(upload_dir(slug, 'kategoriler'), k.resim))
            except: pass
        k.resim = ''
    elif tip == 'banner':
        if k.banner:
            try: os.remove(os.path.join(upload_dir(slug, 'kategoriler'), k.banner))
            except: pass
        k.banner = ''
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/r/<slug>/admin/urun_resim_sil/<int:uid>')
@login_required
def t_urun_resim_sil(slug, tenant, me, uid):
    from flask import jsonify
    u = Urun.query.filter_by(id=uid, tenant_id=tenant.id).first_or_404()
    if u.resim:
        try: os.remove(os.path.join(upload_dir(slug, 'urunler'), u.resim))
        except: pass
    u.resim = ''
    db.session.commit()
    return jsonify({'ok': True})


# ── Sıralama ──
@app.route('/r/<slug>/admin/kat_sirala', methods=['POST'])
@login_required
def t_kat_sirala(slug, tenant, me):
    from flask import jsonify
    ids = request.json.get('ids', [])
    for i, kid in enumerate(ids):
        k = Kategori.query.filter_by(id=kid, tenant_id=tenant.id).first()
        if k: k.sira = i
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/r/<slug>/admin/urun_sirala', methods=['POST'])
@login_required
def t_urun_sirala(slug, tenant, me):
    from flask import jsonify
    ids = request.json.get('ids', [])
    for i, uid in enumerate(ids):
        u = Urun.query.filter_by(id=uid, tenant_id=tenant.id).first()
        if u: u.sira = i
    db.session.commit()
    return jsonify({'ok': True})


if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=5000)
